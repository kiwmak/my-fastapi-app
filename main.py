from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, text
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import os, shutil, uuid, glob, re, logging
from datetime import datetime

# ================= LOGGING =================
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ================= DIRS =================
UPLOAD_DIR = "uploads"
EXPORT_DIR = "exports"
TEMPLATE_DIR = "templates"

for d in [UPLOAD_DIR, EXPORT_DIR, TEMPLATE_DIR]:
    os.makedirs(d, exist_ok=True)

# ================= FASTAPI =================
app = FastAPI(title="Quản lý Test Đốt", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/exports", StaticFiles(directory=EXPORT_DIR), name="exports")

TEMPLATE_FILE = os.path.join(TEMPLATE_DIR, "MAU.xlsx")
LOGO_FILE = os.path.join(TEMPLATE_DIR, "logo.png")

# ================= DB MANAGER =================
class DatabaseManager:
    def __init__(self):
        try:
            db_url = os.environ.get("DATABASE_URL")
            if not db_url:
                raise RuntimeError("DATABASE_URL is not set")

            if db_url.startswith("postgres://"):
                db_url = db_url.replace("postgres://", "postgresql://", 1)

            connect_args = {"sslmode": "require"}

            self.engine = create_engine(
                db_url,
                connect_args=connect_args,
                pool_pre_ping=True,
                echo=False
            )

            logger.info("Connected to PostgreSQL")
            self.ensure_table()

        except Exception as e:
            logger.error(f"DB INIT FAILED: {e}", exc_info=True)
            self.engine = None

    def ensure_table(self):
        if not self.engine:
            return
        query = """
        CREATE TABLE IF NOT EXISTS data (
            "KHÁCH HÀNG" TEXT,
            "ĐƠN HÀNG" TEXT,
            "MÃ HÀNG" TEXT,
            "KÍCH THƯỚC" TEXT,
            "BẤC" TEXT,
            "MÀU" TEXT,
            "HƯƠNG LIỆU" TEXT,
            "NGÀY_TẠO" TIMESTAMP,
            PRIMARY KEY ("ĐƠN HÀNG","MÃ HÀNG")
        )
        """
        with self.engine.begin() as conn:
            conn.execute(text(query))
        logger.info("Table ensured")

    def import_excel(self, path):
        df = pd.read_excel(path, skiprows=1, header=None).fillna("")
        df_out = pd.DataFrame({
            "KHÁCH HÀNG": df.iloc[:, 0],
            "ĐƠN HÀNG": df.iloc[:, 1],
            "MÃ HÀNG": df.iloc[:, 6],
            "KÍCH THƯỚC": df.iloc[:, 12],
            "MÀU": df.iloc[:, 24],
            "HƯƠNG LIỆU": df.iloc[:, 27],
            "BẤC": "",
            "NGÀY_TẠO": datetime.now()
        })

        df_out = df_out[(df_out["ĐƠN HÀNG"] != "") & (df_out["MÃ HÀNG"] != "")]

        sql = """
        INSERT INTO data VALUES
        (:kh,:dh,:mh,:kt,:bac,:mau,:hl,:nt)
        ON CONFLICT ("ĐƠN HÀNG","MÃ HÀNG") DO UPDATE SET
        "KHÁCH HÀNG"=EXCLUDED."KHÁCH HÀNG",
        "KÍCH THƯỚC"=EXCLUDED."KÍCH THƯỚC",
        "MÀU"=EXCLUDED."MÀU",
        "HƯƠNG LIỆU"=EXCLUDED."HƯƠNG LIỆU",
        "BẤC"=EXCLUDED."BẤC",
        "NGÀY_TẠO"=EXCLUDED."NGÀY_TẠO"
        """

        with self.engine.begin() as conn:
            conn.execute(text(sql), [
                dict(
                    kh=r["KHÁCH HÀNG"], dh=r["ĐƠN HÀNG"], mh=r["MÃ HÀNG"],
                    kt=r["KÍCH THƯỚC"], bac=r["BẤC"], mau=r["MÀU"],
                    hl=r["HƯƠNG LIỆU"], nt=r["NGÀY_TẠO"]
                )
                for _, r in df_out.iterrows()
            ])

        return len(df_out)

    def orders(self):
        with self.engine.connect() as c:
            o = c.execute(text('SELECT DISTINCT "ĐƠN HÀNG" FROM data')).fetchall()
            t = c.execute(text('SELECT COUNT(*) FROM data')).scalar()
        return [x[0] for x in o], t

    def order_detail(self, order):
        with self.engine.connect() as c:
            r = c.execute(text('SELECT * FROM data WHERE "ĐƠN HÀNG"=:o'), {"o": order})
            return [dict(x._mapping) for x in r]

db = DatabaseManager()

# ================= EXPORT =================
def export_excel(order, rows):
    wb = load_workbook(TEMPLATE_FILE)
    base = wb.active
    for r in rows:
        ws = wb.copy_worksheet(base)
        ws.title = str(r["MÃ HÀNG"])[:31]
        ws["C5"] = order
        ws["C6"] = r["KHÁCH HÀNG"]
        ws["C7"] = r["HƯƠNG LIỆU"]
        ws["C8"] = r["MÀU"]
        ws["C9"] = r["BẤC"]
        ws["N5"] = r["MÃ HÀNG"]
        ws["N8"] = datetime.now().strftime("%Y-%m-%d")
        if os.path.exists(LOGO_FILE):
            ws.add_image(OpenpyxlImage(LOGO_FILE), "A1")

    wb.remove(base)
    name = f"{order}_{uuid.uuid4().hex[:6]}.xlsx"
    path = os.path.join(EXPORT_DIR, name)
    wb.save(path)
    return name

# ================= API =================
@app.get("/", response_class=HTMLResponse)
def home():
    return "<h3>API OK</h3>"

@app.post("/api/import")
async def api_import(file: UploadFile = File(...)):
    path = f"{UPLOAD_DIR}/{uuid.uuid4()}_{file.filename}"
    with open(path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    count = db.import_excel(path)
    os.remove(path)
    return {"success": True, "rows": count}

@app.get("/api/orders")
def api_orders():
    o, t = db.orders()
    return {"orders": o, "total_rows": t}

@app.get("/api/order/{order}")
def api_order(order: str):
    data = db.order_detail(order)
    if not data:
        raise HTTPException(404, "Not found")
    return {"data": data}

@app.post("/api/export/{order}")
def api_export(order: str):
    rows = db.order_detail(order)
    name = export_excel(order, rows)
    return {"download": f"/exports/{name}"}

@app.get("/api/download/{file}")
def download(file: str):
    return FileResponse(os.path.join(EXPORT_DIR, file))

# ================= RUN =================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000)
