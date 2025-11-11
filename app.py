from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import shutil
from datetime import datetime
import uuid
import logging
import re
import glob

# Cấu hình logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Quản lý Test Đốt", version="1.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Thư mục
UPLOAD_DIR = "uploads"
EXPORT_DIR = "exports"
DATA_DIR = "data"
TEMPLATE_DIR = "templates"

for directory in [UPLOAD_DIR, EXPORT_DIR, DATA_DIR, TEMPLATE_DIR]:
    os.makedirs(directory, exist_ok=True)

# File paths
DATA_FILE = os.path.join(DATA_DIR, "data.xlsx")
TEMPLATE_FILE = os.path.join(TEMPLATE_DIR, "MAU.xlsx")
LOGO_FILE = os.path.join(TEMPLATE_DIR, "logo.png")


class DataManager:
    def __init__(self):
        self.data_file = DATA_FILE
        self.ensure_data_file()

    def ensure_data_file(self):
        if not os.path.exists(self.data_file):
            df = pd.DataFrame(columns=[
                "KHÁCH HÀNG", "ĐƠN HÀNG", "MÃ HÀNG", "KÍCH THƯỚC",
                "BẤC", "MÀU", "HƯƠNG LIỆU", "NGÀY_TẠO"
            ])
            self.save_data(df)
            logger.info("Đã tạo file data mới")

    def load_data(self):
        try:
            if os.path.exists(self.data_file):
                df = pd.read_excel(self.data_file)
                df = df.fillna("")
                logger.info(f"Đã load {len(df)} dòng dữ liệu")
                return df
            return pd.DataFrame()
        except Exception as e:
            logger.error(f"Lỗi đọc data: {e}")
            return pd.DataFrame()

    def save_data(self, df):
        try:
            df_cleaned = df.fillna("")
            df_cleaned.to_excel(self.data_file, index=False)
            return True
        except Exception as e:
            logger.error(f"Lỗi lưu data: {e}")
            return False

    def import_data(self, file_path):
        try:
            df_new = pd.read_excel(file_path).fillna("")
            df_existing = self.load_data()

            # Map các tên cột
            column_mapping = {
                "KHÁCH HÀNG": ["KHÁCH HÀNG", "CUSTOMER", "TEN_KHACH_HANG"],
                "ĐƠN HÀNG": ["ĐƠN HÀNG", "ORDER", "MA_DON_HANG"],
                "MÃ HÀNG": ["MÃ HÀNG", "PRODUCT_CODE", "MA_HANG"],
                "KÍCH THƯỚC": ["KÍCH THƯỚC", "SIZE", "KICH_THUOC"],
                "BẤC": ["BẤC", "WICK", "Bac"],
                "MÀU": ["MÀU", "COLOR", "MAU"],
                "HƯƠNG LIỆU": ["HƯƠNG LIỆU", "FRAGRANCE", "HUONG_LIEU"]
            }

            for standard_name, possible_names in column_mapping.items():
                for possible_name in possible_names:
                    if possible_name in df_new.columns:
                        df_new = df_new.rename(columns={possible_name: standard_name})
                        break

            required_cols = ["KHÁCH HÀNG", "ĐƠN HÀNG", "MÃ HÀNG"]
            missing_cols = [col for col in required_cols if col not in df_new.columns]

            if missing_cols:
                return {"success": False, "message": f"Thiếu cột: {', '.join(missing_cols)}"}

            df_new["NGÀY_TẠO"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if not df_existing.empty:
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            else:
                df_combined = df_new

            df_combined = df_combined.drop_duplicates(subset=["ĐƠN HÀNG", "MÃ HÀNG"], keep='last')

            if self.save_data(df_combined):
                return {
                    "success": True,
                    "message": f"Import thành công: {len(df_new)} dòng",
                    "total_rows": len(df_combined)
                }
            else:
                return {"success": False, "message": "Lỗi khi lưu dữ liệu"}

        except Exception as e:
            logger.error(f"Lỗi import: {e}")
            return {"success": False, "message": f"Lỗi: {str(e)}"}


class ExportManager:
    def __init__(self):
        self.template_file = TEMPLATE_FILE
        self.logo_file = LOGO_FILE

    def get_reports_list(self):
        """Lấy danh sách các báo cáo đã tạo"""
        try:
            reports = []
            if os.path.exists(EXPORT_DIR):
                for file_path in glob.glob(os.path.join(EXPORT_DIR, "*.xlsx")):
                    file_name = os.path.basename(file_path)
                    file_size = os.path.getsize(file_path)
                    created_time = datetime.fromtimestamp(os.path.getctime(file_path))

                    # Phân tích tên file để lấy thông tin
                    parts = file_name.replace('.xlsx', '').split('_')
                    order_no = parts[0] if parts else "Unknown"

                    reports.append({
                        "filename": file_name,
                        "order_no": order_no,
                        "file_size": file_size,
                        "created_time": created_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "file_path": file_path
                    })

            # Sắp xếp theo thời gian tạo mới nhất
            reports.sort(key=lambda x: x["created_time"], reverse=True)
            return reports
        except Exception as e:
            logger.error(f"Lỗi lấy danh sách báo cáo: {e}")
            return []

    def delete_report(self, filename: str):
        """Xóa báo cáo"""
        try:
            file_path = os.path.join(EXPORT_DIR, filename)
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Đã xóa báo cáo: {filename}")
                return True
            return False
        except Exception as e:
            logger.error(f"Lỗi xóa báo cáo {filename}: {e}")
            return False

    def export_with_template(self, order_no: str, order_data: pd.DataFrame, output_path: str) -> dict:
        """Xuất báo cáo theo template MAU.xlsx"""
        try:
            if not os.path.exists(self.template_file):
                return {"success": False,
                        "message": f"Không tìm thấy file template: {os.path.basename(self.template_file)}"}

            logger.info(f"Bắt đầu xuất báo cáo cho đơn hàng: {order_no}")

            shutil.copy2(self.template_file, output_path)
            wb = load_workbook(output_path)
            template_sheet = wb.worksheets[0]
            template_sheet_name = template_sheet.title

            sheets_created = 0
            ma_hang_list = order_data["MÃ HÀNG"].dropna().unique()

            for ma_hang in ma_hang_list:
                ma_hang_str = str(ma_hang).strip()
                if not ma_hang_str:
                    continue

                try:
                    new_sheet = wb.copy_worksheet(template_sheet)
                    sheet_name = ma_hang_str[:31]
                    original_name = sheet_name
                    counter = 1

                    while sheet_name in wb.sheetnames:
                        sheet_name = f"{original_name}_{counter}"
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        counter += 1

                    new_sheet.title = sheet_name
                    sheets_created += 1

                    product_data = order_data[order_data["MÃ HÀNG"].astype(str) == ma_hang_str]
                    if product_data.empty:
                        continue

                    row_data = product_data.iloc[0]
                    mapping = self.get_cell_mapping(order_no, row_data)

                    for cell_ref, value in mapping.items():
                        try:
                            if value not in (None, ""):
                                new_sheet[cell_ref] = value
                        except Exception as e:
                            logger.warning(f"Không thể ghi ô {cell_ref}: {e}")

                    self.insert_logo(new_sheet)
                    logger.info(f"Đã tạo sheet cho mã hàng: {ma_hang_str}")

                except Exception as e:
                    logger.error(f"Lỗi tạo sheet cho {ma_hang}: {e}")
                    continue

            if sheets_created > 0 and template_sheet_name in wb.sheetnames:
                try:
                    wb.remove(wb[template_sheet_name])
                except Exception as e:
                    logger.warning(f"Không thể xóa sheet template: {e}")

            wb.save(output_path)
            logger.info(f"Xuất báo cáo thành công: {sheets_created} sheets")

            return {
                "success": True,
                "message": f"Xuất báo cáo thành công: {sheets_created} mã hàng",
                "sheets_created": sheets_created,
                "file_path": output_path
            }

        except Exception as e:
            logger.error(f"Lỗi xuất báo cáo: {e}")
            return {"success": False, "message": f"Lỗi xuất báo cáo: {str(e)}"}

    def get_cell_mapping(self, order_no: str, row_data: pd.Series) -> dict:
        """Mapping dữ liệu vào các ô trong template"""
        mapping = {
            "C5": order_no,  # Đơn hàng
            "C6": row_data.get("KHÁCH HÀNG", ""),  # Khách hàng
            "C7": row_data.get("HƯƠNG LIỆU", ""),  # Hương liệu
            "C8": row_data.get("MÀU", ""),  # Màu
            "C9": row_data.get("BẤC", ""),  # Bấc
            "N5": row_data.get("MÃ HÀNG", ""),  # Mã hàng
            "N8": datetime.now().strftime("%Y-%m-%d")  # Ngày test
        }

        kich_thuoc = str(row_data.get("KÍCH THƯỚC", ""))
        duong_kinh, chieu_cao = self.parse_kich_thuoc(kich_thuoc)

        mapping["N6"] = duong_kinh  # Đường kính
        mapping["S6"] = chieu_cao  # Chiều cao

        return mapping

    def parse_kich_thuoc(self, kich_thuoc: str) -> tuple:
        """Phân tích chuỗi kích thước thành đường kính và chiều cao"""
        if not kich_thuoc:
            return "", ""

        try:
            kich_thuoc = str(kich_thuoc).lower().replace('×', 'x').replace('*', 'x').replace(' ', '')
            parts = re.findall(r'[\d.]+', kich_thuoc)

            if len(parts) >= 2:
                duong_kinh = float(parts[0])
                chieu_cao = float(parts[1])

                if 'cm' in kich_thuoc:
                    duong_kinh *= 10
                    chieu_cao *= 10

                return round(duong_kinh, 1), round(chieu_cao, 1)
            else:
                return "", ""

        except Exception as e:
            logger.warning(f"Lỗi phân tích kích thước '{kich_thuoc}': {e}")
            return "", ""

    def insert_logo(self, worksheet):
        """Chèn logo vào worksheet"""
        try:
            if os.path.exists(self.logo_file):
                img = OpenpyxlImage(self.logo_file)
                worksheet.add_image(img, 'A1')
                logger.info("Đã chèn logo vào báo cáo")
        except Exception as e:
            logger.warning(f"Không thể chèn logo: {e}")


data_manager = DataManager()
export_manager = ExportManager()

# HTML Template với danh sách báo cáo
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Quản lý Test Đốt</title>
    <script src="https://cdn.jsdelivr.net/npm/axios/dist/axios.min.js"></script>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f8f9fa; }
        .container { max-width: 1400px; }
        .card { border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); border: none; }
        .btn-export { background: linear-gradient(135deg, #28a745, #20c997); color: white; border: none; }
        .btn-export:hover { transform: translateY(-1px); box-shadow: 0 4px 8px rgba(40, 167, 69, 0.3); }
        .report-item { transition: all 0.3s ease; border-left: 4px solid #007bff; }
        .report-item:hover { transform: translateX(5px); background-color: #f8f9fa; }
        .nav-tabs .nav-link.active { font-weight: 600; border-bottom: 3px solid #007bff; }
        .file-size { font-size: 0.85rem; color: #6c757d; }
        .action-buttons .btn { padding: 0.25rem 0.5rem; font-size: 0.875rem; }
    </style>
</head>
<body>
    <div class="container mt-4">
        <h1 class="text-center mb-4">🏭 Quản lý Test Đốt</h1>

        <!-- Navigation Tabs -->
        <ul class="nav nav-tabs mb-4" id="mainTabs" role="tablist">
            <li class="nav-item" role="presentation">
                <button class="nav-link active" id="dashboard-tab" data-bs-toggle="tab" data-bs-target="#dashboard" type="button" role="tab">
                    <i class="fas fa-tachometer-alt me-2"></i>Dashboard
                </button>
            </li>
            <li class="nav-item" role="presentation">
                <button class="nav-link" id="reports-tab" data-bs-toggle="tab" data-bs-target="#reports" type="button" role="tab">
                    <i class="fas fa-file-alt me-2"></i>Báo cáo đã tạo
                </button>
            </li>
            <li class="nav-item" role="presentation">
                <button class="nav-link" id="templates-tab" data-bs-toggle="tab" data-bs-target="#templates" type="button" role="tab">
                    <i class="fas fa-cog me-2"></i>Quản lý Template
                </button>
            </li>
        </ul>

        <!-- Tab Content -->
        <div class="tab-content" id="mainTabsContent">
            <!-- Dashboard Tab -->
            <div class="tab-pane fade show active" id="dashboard" role="tabpanel">
                <div class="row">
                    <div class="col-md-6">
                        <div class="card p-3 mb-4">
                            <h5><i class="fas fa-file-import me-2"></i>Import Dữ liệu</h5>
                            <input type="file" id="fileInput" class="form-control mb-2" accept=".xlsx">
                            <button class="btn btn-primary" onclick="importData()">
                                <i class="fas fa-upload me-2"></i>Import Excel
                            </button>
                            <div id="importResult" class="mt-2"></div>
                        </div>
                    </div>

                    <div class="col-md-6">
                        <div class="card p-3">
                            <h5><i class="fas fa-chart-bar me-2"></i>Thống kê</h5>
                            <p><i class="fas fa-boxes me-2"></i>Tổng đơn hàng: <span id="totalOrders" class="fw-bold">0</span></p>
                            <p><i class="fas fa-database me-2"></i>Tổng dòng dữ liệu: <span id="totalRows" class="fw-bold">0</span></p>
                            <button class="btn btn-info" onclick="loadStats()">
                                <i class="fas fa-sync-alt me-2"></i>Làm mới
                            </button>
                        </div>
                    </div>
                </div>

                <div class="card p-3 mt-4">
                    <h5><i class="fas fa-rocket me-2"></i>Xuất báo cáo theo template</h5>
                    <div class="row">
                        <div class="col-md-6">
                            <select id="orderSelect" class="form-select mb-2" onchange="loadOrderDetail()">
                                <option value="">-- Chọn đơn hàng --</option>
                            </select>
                        </div>
                        <div class="col-md-6">
                            <button class="btn btn-export w-100" onclick="exportWithTemplate()">
                                <i class="fas fa-file-export me-2"></i>Xuất báo cáo theo mẫu
                            </button>
                        </div>
                    </div>
                    <div id="orderDetail" class="mt-3"></div>
                </div>
            </div>

            <!-- Reports Tab -->
            <div class="tab-pane fade" id="reports" role="tabpanel">
                <div class="card p-3">
                    <div class="d-flex justify-content-between align-items-center mb-3">
                        <h5><i class="fas fa-history me-2"></i>Danh sách báo cáo đã tạo</h5>
                        <button class="btn btn-outline-primary" onclick="loadReports()">
                            <i class="fas fa-sync-alt me-2"></i>Làm mới
                        </button>
                    </div>

                    <div class="table-responsive">
                        <table class="table table-hover">
                            <thead class="table-light">
                                <tr>
                                    <th>Tên file</th>
                                    <th>Đơn hàng</th>
                                    <th>Kích thước</th>
                                    <th>Ngày tạo</th>
                                    <th>Thao tác</th>
                                </tr>
                            </thead>
                            <tbody id="reportsList">
                                <tr>
                                    <td colspan="5" class="text-center text-muted py-4">
                                        <i class="fas fa-spinner fa-spin me-2"></i>Đang tải...
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                    </div>

                    <div class="d-flex justify-content-between align-items-center mt-3">
                        <small class="text-muted" id="reportsCount">Đang tải...</small>
                        <button class="btn btn-outline-danger btn-sm" onclick="clearAllReports()">
                            <i class="fas fa-trash me-2"></i>Xóa tất cả
                        </button>
                    </div>
                </div>
            </div>

            <!-- Templates Tab -->
            <div class="tab-pane fade" id="templates" role="tabpanel">
                <div class="card p-3">
                    <h5><i class="fas fa-cog me-2"></i>Quản lý Template</h5>
                    <div class="row">
                        <div class="col-md-6">
                            <div class="mb-3">
                                <label class="form-label">File template (MAU.xlsx):</label>
                                <input type="file" id="templateFile" class="form-control" accept=".xlsx">
                                <small class="form-text text-muted">Tải lên file template mới</small>
                            </div>
                            <button class="btn btn-warning" onclick="uploadTemplate()">
                                <i class="fas fa-upload me-2"></i>Tải lên Template
                            </button>
                        </div>
                        <div class="col-md-6">
                            <div class="mb-3">
                                <label class="form-label">Logo:</label>
                                <input type="file" id="logoFile" class="form-control" accept=".png,.jpg,.jpeg">
                                <small class="form-text text-muted">Tải lên logo mới (PNG/JPG)</small>
                            </div>
                            <button class="btn btn-warning" onclick="uploadLogo()">
                                <i class="fas fa-image me-2"></i>Tải lên Logo
                            </button>
                        </div>
                    </div>
                    <div id="templateResult" class="mt-2"></div>
                </div>
            </div>
        </div>
    </div>

    <!-- Bootstrap JS -->
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>

    <script>
        const API_BASE = '/api';

        // Khởi tạo khi trang load
        document.addEventListener('DOMContentLoaded', function() {
            loadStats();
            loadOrders();
            loadReports();
        });

        async function loadStats() {
            try {
                const response = await axios.get(`${API_BASE}/orders`);
                document.getElementById('totalOrders').textContent = response.data.orders.length;
                if (response.data.orders.length > 0) {
                    const detail = await axios.get(`${API_BASE}/order/${response.data.orders[0]}`);
                    document.getElementById('totalRows').textContent = detail.data.total_items;
                }
            } catch (error) {
                console.error('Lỗi:', error);
            }
        }

        async function importData() {
            const file = document.getElementById('fileInput').files[0];
            if (!file) {
                showAlert('importResult', false, 'Vui lòng chọn file trước!');
                return;
            }

            const formData = new FormData();
            formData.append('file', file);

            try {
                const response = await axios.post(`${API_BASE}/import`, formData);
                showAlert('importResult', response.data.success, response.data.message);
                loadStats();
                loadOrders();
            } catch (error) {
                showAlert('importResult', false, 'Lỗi kết nối');
            }
        }

        async function loadOrders() {
            try {
                const response = await axios.get(`${API_BASE}/orders`);
                const select = document.getElementById('orderSelect');
                select.innerHTML = '<option value="">-- Chọn đơn hàng --</option>';
                response.data.orders.forEach(order => {
                    const option = document.createElement('option');
                    option.value = order;
                    option.textContent = order;
                    select.appendChild(option);
                });
            } catch (error) {
                console.error('Lỗi:', error);
            }
        }

        async function loadOrderDetail() {
            const orderNo = document.getElementById('orderSelect').value;
            if (!orderNo) return;

            try {
                const response = await axios.get(`${API_BASE}/order/${orderNo}`);
                let html = '<div class="table-responsive"><table class="table table-striped"><thead><tr>';
                if (response.data.data.length > 0) {
                    Object.keys(response.data.data[0]).forEach(key => {
                        html += `<th>${key}</th>`;
                    });
                    html += '</tr></thead><tbody>';
                    response.data.data.forEach(row => {
                        html += '<tr>';
                        Object.values(row).forEach(value => {
                            html += `<td>${value || ''}</td>`;
                        });
                        html += '</tr>';
                    });
                    html += '</tbody></table></div>';
                }
                document.getElementById('orderDetail').innerHTML = html;
            } catch (error) {
                console.error('Lỗi:', error);
            }
        }

        async function exportWithTemplate() {
            const orderNo = document.getElementById('orderSelect').value;
            if (!orderNo) {
                showAlert('orderDetail', false, 'Vui lòng chọn đơn hàng!');
                return;
            }

            try {
                const response = await axios.post(`${API_BASE}/export-template/${orderNo}`);
                if (response.data.success) {
                    showAlert('orderDetail', true, response.data.message);
                    // Tự động download file
                    const downloadLink = document.createElement('a');
                    downloadLink.href = response.data.download_url;
                    downloadLink.download = '';
                    document.body.appendChild(downloadLink);
                    downloadLink.click();
                    document.body.removeChild(downloadLink);

                    // Làm mới danh sách báo cáo
                    setTimeout(loadReports, 1000);
                } else {
                    showAlert('orderDetail', false, response.data.message);
                }
            } catch (error) {
                showAlert('orderDetail', false, 'Lỗi xuất báo cáo');
            }
        }

        async function loadReports() {
            try {
                const response = await axios.get(`${API_BASE}/reports`);
                const reports = response.data.reports || [];
                const tbody = document.getElementById('reportsList');

                if (reports.length === 0) {
                    tbody.innerHTML = `
                        <tr>
                            <td colspan="5" class="text-center text-muted py-4">
                                <i class="fas fa-inbox me-2"></i>Chưa có báo cáo nào được tạo
                            </td>
                        </tr>
                    `;
                    document.getElementById('reportsCount').textContent = '0 báo cáo';
                    return;
                }

                let html = '';
                reports.forEach(report => {
                    const fileSize = (report.file_size / 1024).toFixed(1) + ' KB';
                    html += `
                        <tr class="report-item">
                            <td>
                                <i class="fas fa-file-excel text-success me-2"></i>
                                <strong>${report.filename}</strong>
                            </td>
                            <td>${report.order_no}</td>
                            <td><span class="file-size">${fileSize}</span></td>
                            <td>${report.created_time}</td>
                            <td class="action-buttons">
                                <button class="btn btn-success btn-sm me-1" onclick="downloadReport('${report.filename}')" title="Tải xuống">
                                    <i class="fas fa-download"></i>
                                </button>
                                <button class="btn btn-primary btn-sm me-1" onclick="viewReport('${report.filename}')" title="Xem trước">
                                    <i class="fas fa-eye"></i>
                                </button>
                                <button class="btn btn-danger btn-sm" onclick="deleteReport('${report.filename}')" title="Xóa">
                                    <i class="fas fa-trash"></i>
                                </button>
                            </td>
                        </tr>
                    `;
                });

                tbody.innerHTML = html;
                document.getElementById('reportsCount').textContent = `${reports.length} báo cáo`;

            } catch (error) {
                console.error('Lỗi tải danh sách báo cáo:', error);
                document.getElementById('reportsList').innerHTML = `
                    <tr>
                        <td colspan="5" class="text-center text-danger py-4">
                            <i class="fas fa-exclamation-triangle me-2"></i>Lỗi tải danh sách báo cáo
                        </td>
                    </tr>
                `;
            }
        }

        async function downloadReport(filename) {
            try {
                const downloadLink = document.createElement('a');
                downloadLink.href = `${API_BASE}/download/${filename}`;
                downloadLink.download = filename;
                document.body.appendChild(downloadLink);
                downloadLink.click();
                document.body.removeChild(downloadLink);
            } catch (error) {
                alert('Lỗi tải file: ' + error);
            }
        }

        async function viewReport(filename) {
            // Mở file trong tab mới (chỉ xem được nếu browser hỗ trợ)
            window.open(`${API_BASE}/download/${filename}`, '_blank');
        }

        async function deleteReport(filename) {
            if (!confirm(`Bạn có chắc muốn xóa báo cáo "${filename}"?`)) {
                return;
            }

            try {
                const response = await axios.delete(`${API_BASE}/reports/${filename}`);
                if (response.data.success) {
                    showAlert('reportsList', true, 'Đã xóa báo cáo thành công');
                    loadReports();
                } else {
                    showAlert('reportsList', false, response.data.message);
                }
            } catch (error) {
                showAlert('reportsList', false, 'Lỗi xóa báo cáo');
            }
        }

        async function clearAllReports() {
            if (!confirm('Bạn có chắc muốn xóa TẤT CẢ báo cáo? Hành động này không thể hoàn tác!')) {
                return;
            }

            try {
                const response = await axios.delete(`${API_BASE}/reports`);
                if (response.data.success) {
                    showAlert('reportsList', true, 'Đã xóa tất cả báo cáo thành công');
                    loadReports();
                } else {
                    showAlert('reportsList', false, response.data.message);
                }
            } catch (error) {
                showAlert('reportsList', false, 'Lỗi xóa báo cáo');
            }
        }

        async function uploadTemplate() {
            const file = document.getElementById('templateFile').files[0];
            if (!file) {
                showAlert('templateResult', false, 'Vui lòng chọn file template!');
                return;
            }

            const formData = new FormData();
            formData.append('file', file);

            try {
                const response = await axios.post(`${API_BASE}/upload-template`, formData);
                showAlert('templateResult', response.data.success, response.data.message);
            } catch (error) {
                showAlert('templateResult', false, 'Lỗi tải lên template');
            }
        }

        async function uploadLogo() {
            const file = document.getElementById('logoFile').files[0];
            if (!file) {
                showAlert('templateResult', false, 'Vui lòng chọn file logo!');
                return;
            }

            const formData = new FormData();
            formData.append('file', file);

            try {
                const response = await axios.post(`${API_BASE}/upload-logo`, formData);
                showAlert('templateResult', response.data.success, response.data.message);
            } catch (error) {
                showAlert('templateResult', false, 'Lỗi tải lên logo');
            }
        }

        function showAlert(containerId, success, message) {
            const alertClass = success ? 'alert-success' : 'alert-danger';
            const icon = success ? '✅' : '❌';
            const alertDiv = document.createElement('div');
            alertDiv.className = `alert ${alertClass} alert-dismissible fade show`;
            alertDiv.innerHTML = `
                ${icon} ${message}
                <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
            `;

            const container = document.getElementById(containerId);
            // Xóa alert cũ
            const oldAlert = container.querySelector('.alert');
            if (oldAlert) {
                oldAlert.remove();
            }

            container.appendChild(alertDiv);

            // Tự động xóa sau 5 giây
            setTimeout(() => {
                if (alertDiv.parentNode) {
                    alertDiv.remove();
                }
            }, 5000);
        }
    </script>
</body>
</html>
"""


# API Routes
@app.get("/")
async def root():
    return HTMLResponse(HTML_TEMPLATE)


@app.get("/api/orders")
async def get_orders():
    df = data_manager.load_data()
    orders = df["ĐƠN HÀNG"].dropna().unique().tolist() if not df.empty else []
    return {"orders": sorted(orders)}


@app.get("/api/order/{order_no}")
async def get_order(order_no: str):
    df = data_manager.load_data()
    order_data = df[df["ĐƠN HÀNG"].astype(str) == order_no] if not df.empty else []
    return {
        "data": order_data.fillna("").to_dict('records'),
        "total_items": len(order_data)
    }


@app.post("/api/import")
async def import_data(file: UploadFile = File(...)):
    try:
        file_path = os.path.join(UPLOAD_DIR, f"temp_{uuid.uuid4().hex}.xlsx")
        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        result = data_manager.import_data(file_path)

        try:
            os.remove(file_path)
        except:
            pass

        return result
    except Exception as e:
        return {"success": False, "message": f"Lỗi: {str(e)}"}


@app.get("/api/reports")
async def get_reports():
    """Lấy danh sách các báo cáo đã tạo"""
    reports = export_manager.get_reports_list()
    return {"reports": reports}


@app.delete("/api/reports/{filename}")
async def delete_report(filename: str):
    """Xóa một báo cáo"""
    success = export_manager.delete_report(filename)
    if success:
        return {"success": True, "message": "Đã xóa báo cáo thành công"}
    else:
        return {"success": False, "message": "Không thể xóa báo cáo"}


@app.delete("/api/reports")
async def clear_all_reports():
    """Xóa tất cả báo cáo"""
    try:
        reports = export_manager.get_reports_list()
        deleted_count = 0

        for report in reports:
            if export_manager.delete_report(report["filename"]):
                deleted_count += 1

        return {
            "success": True,
            "message": f"Đã xóa {deleted_count} báo cáo thành công",
            "deleted_count": deleted_count
        }
    except Exception as e:
        return {"success": False, "message": f"Lỗi xóa báo cáo: {str(e)}"}


@app.post("/api/export-template/{order_no}")
async def export_with_template(order_no: str):
    """Xuất báo cáo theo template"""
    try:
        df = data_manager.load_data()
        order_data = df[df["ĐƠN HÀNG"].astype(str) == order_no]

        if order_data.empty:
            return {"success": False, "message": "Không có dữ liệu đơn hàng"}

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_filename = f"{order_no}_BAO_CAO_{timestamp}.xlsx"
        export_path = os.path.join(EXPORT_DIR, export_filename)

        result = export_manager.export_with_template(order_no, order_data, export_path)

        if result["success"]:
            result["download_url"] = f"/api/download/{export_filename}"
            return result
        else:
            return result

    except Exception as e:
        logger.error(f"Lỗi xuất báo cáo template: {e}")
        return {"success": False, "message": f"Lỗi xuất báo cáo: {str(e)}"}


@app.post("/api/upload-template")
async def upload_template(file: UploadFile = File(...)):
    """Tải lên file template mới"""
    try:
        if not file.filename.endswith('.xlsx'):
            return {"success": False, "message": "Chỉ chấp nhận file .xlsx"}

        file_path = os.path.join(TEMPLATE_DIR, "MAU.xlsx")
        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        return {"success": True, "message": "Đã cập nhật template thành công"}
    except Exception as e:
        return {"success": False, "message": f"Lỗi tải lên template: {str(e)}"}


@app.post("/api/upload-logo")
async def upload_logo(file: UploadFile = File(...)):
    """Tải lên logo mới"""
    try:
        if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
            return {"success": False, "message": "Chỉ chấp nhận file PNG, JPG, JPEG"}

        file_path = os.path.join(TEMPLATE_DIR, "logo.png")
        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        return {"success": True, "message": "Đã cập nhật logo thành công"}
    except Exception as e:
        return {"success": False, "message": f"Lỗi tải lên logo: {str(e)}"}


@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """Download file"""
    file_path = os.path.join(EXPORT_DIR, filename)
    if os.path.exists(file_path):
        return FileResponse(
            file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        raise HTTPException(status_code=404, detail="File không tồn tại")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8080, log_level="info")